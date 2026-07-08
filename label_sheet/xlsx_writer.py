"""
Multi-sheet workbook writer (openpyxl).

Writes one worksheet per year using the canonical schema, with:
- an embedded thumbnail in column A of every data row,
- colored header bands per column group,
- a frozen header row,
- Excel data-validation dropdowns on controlled-vocabulary columns (including
  the expert confirm/reject/update columns),
- a cell comment on the first header cell of each model group recording exactly
  which model / weights / feature extractor produced those estimates.

Every worksheet has identical columns even when a year lacks data, satisfying
the "same formatting across all years" requirement.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .config import ModelConfig, run_timestamp
from .schema import COLUMNS, DATA_KEYS, GROUP_FILLS, Column
from .thumbnails import THUMB_SIZE, get_thumbnail

HEADER_FONT = Font(bold=True, size=10)
ROW_HEIGHT = 74  # points; fits a 96px thumbnail with margin
THUMB_COL_WIDTH = 14


def _dropdown_formula(options: Sequence[str]) -> str:
    joined = ",".join(o for o in options if o != "")
    return f'"{joined}"'


def _model_comment_text(models: Sequence[ModelConfig], model_id: str) -> str:
    cfg = next((m for m in models if m.model_id == model_id), None)
    if cfg is None:
        return f"Estimates from model {model_id}."
    return (
        f"Estimated by: {cfg.display_name}\n"
        f"model_id: {cfg.model_id}\n"
        f"weights: models/{cfg.weights}\n"
        f"features: {cfg.feature_extractor}\n"
        f"bootstrap: {cfg.bootstrap}\n"
        f"generated: {run_timestamp()}\n"
        f"{cfg.notes}\n\n"
        f"'*_in_training_set' flags rows whose source year likely appeared in "
        f"this model's training data (trained_on={list(cfg.trained_on)}); treat "
        f"those estimates as optimistic."
    )


def _write_sheet(ws, df: pd.DataFrame, models: Sequence[ModelConfig]) -> None:
    # Column widths + header row.
    for col_idx, col in enumerate(COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = (
            THUMB_COL_WIDTH if col.is_image else col.width
        )
        cell = ws.cell(row=1, column=col_idx, value=col.header)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        fill_argb = GROUP_FILLS.get(col.group)
        if fill_argb:
            cell.fill = PatternFill("solid", fgColor=fill_argb)

    # Model-provenance comments on the first header cell of each model group.
    for model_id in ("m1", "m2"):
        first = next((i for i, c in enumerate(COLUMNS, start=1)
                      if c.group == f"model_{model_id}"), None)
        if first is not None:
            ws.cell(row=1, column=first).comment = Comment(
                _model_comment_text(models, model_id), "label_sheet")

    # Freeze the header row AND column A (the thumbnail), so the photo stays
    # visible next to the far-right expert_ columns while the reviewer scrolls.
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30

    # Data rows.
    for r, (_, record) in enumerate(df.iterrows(), start=2):
        ws.row_dimensions[r].height = ROW_HEIGHT
        for col_idx, col in enumerate(COLUMNS, start=1):
            if col.is_image:
                continue
            value = record.get(col.key, "")
            cell = ws.cell(row=r, column=col_idx, value=value if value != "" else None)
            if col.wrap:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center")

        # Thumbnail in column A.
        abs_path = record.get("abs_path", "")
        if abs_path:
            thumb = get_thumbnail(abs_path)
            if thumb is not None:
                img = XLImage(str(thumb))
                img.width = THUMB_SIZE
                img.height = THUMB_SIZE
                ws.add_image(img, f"A{r}")

    _add_dropdowns(ws, len(df))


def _add_dropdowns(ws, n_rows: int) -> None:
    if n_rows <= 0:
        return
    last = n_rows + 1
    for col_idx, col in enumerate(COLUMNS, start=1):
        if not col.dropdown:
            continue
        letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=_dropdown_formula(col.dropdown),
            allow_blank=True,
            showDropDown=False,  # False => arrow IS shown (Excel quirk)
        )
        dv.error = "Pick a value from the list."
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{last}")


def _write_readme(wb, models: Sequence[ModelConfig]) -> None:
    """Write a human-facing 'README' instructions tab as the first sheet.

    Self-documents the review workflow so the shared file makes sense without the
    accompanying email. Plain text only, so it survives an Excel -> Google Sheets
    import unchanged.
    """
    ws = wb.create_sheet(title="README")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 96
    title = Font(bold=True, size=14)
    head = Font(bold=True, size=11)
    wrap = Alignment(vertical="top", wrap_text=True)

    counter = {"r": 0}

    def row(a="", b="", *, font=None, fill=None):
        counter["r"] += 1
        r = counter["r"]
        ca, cb = ws.cell(r, 1, a), ws.cell(r, 2, b)
        ca.alignment = wrap
        cb.alignment = wrap
        if font:
            ca.font = font
        if fill:
            ca.fill = PatternFill("solid", fgColor=fill)
        return r

    m1 = next((m for m in models if m.model_id == "m1"), None)
    m2 = next((m for m in models if m.model_id == "m2"), None)

    row("Green Crab Molt Label Sheet", font=title)
    row("", f"Generated {run_timestamp()}. One worksheet per study year "
            "(2016, 2017, 2018, 2019_crate, 2026). Each row is ONE crab image, "
            "except image-less paper-notebook observation rows on the 2026 tab.")
    row()
    row("HOW TO REVIEW", font=head, fill="FFEAD1DC")
    row("", "For each row, judge the AI estimates against the photo and fill the "
            "plum 'expert_*' columns on the right. Do NOT edit the source or model "
            "columns — put your corrections in the expert_* columns so we keep the "
            "original alongside your verdict.")
    row("1. expert_status", "Pick from the dropdown: confirm (AI looks right), "
                            "reject (AI is wrong; leave overrides blank or explain), "
                            "update (AI is close but needs the corrected value).")
    row("2. expert_days_to_molt", "Your best estimate of days until molt, if known.")
    row("3. expert_molt_phase", "Dropdown: intermolt / pre_molt / peeler_imminent / "
                               "molted / dead / unknown.")
    row("4. expert_sex", "Dropdown: male / female / unknown.")
    row("5. expert_view", "Dropdown: ventral / dorsal / side / unknown.")
    row("6. expert_molt_indicators", "Free text: shell color/translucency, seam/suture, "
                                     "side cracking, halo, dusky, etc.")
    row("7. expert_notes", "Anything else — disagreements, image problems, "
                          "'wrong crab', 'restocked cell', etc.")
    row()
    row("COLUMN GROUPS (by color)", font=head, fill="FFEFEFEF")
    row("Gray - identity/paths", "year, dataset, crab_id, cell, image path + the "
                                "thumbnail in column A.")
    row("Blue - ground truth", "Labels extracted from the original records (folder "
                              "names, the year workbooks, the crate .docx, the 2026 "
                              "condo sheet + filenames). Blank where the source had none.")
    row("Green - aux metadata", "condo_id, cell_id, degree_of_molt, is_in_situ.")
    row("Orange - ViT (Model A)", (m1.display_name if m1 else "primary estimator")
        + ". AI estimate of days_to_molt + phase/sex/view/molt indicators, in the "
          "'ViT_' columns. Hover the 'ViT_days_to_molt' header for the exact model "
          "+ weights.")
    row("Yellow - OpenCLIP (Model B)", (m2.display_name if m2 else "bootstrap estimator")
        + ". A second, independent AI estimate in the 'OpenCLIP_' columns, with "
          "OpenCLIP_days_to_molt_std as an uncertainty. Hover the "
          "'OpenCLIP_days_to_molt' header for details.")
    row("Plum - expert review", "Your columns (above). This is the point of the sheet.")
    row()
    row("IMPORTANT CAVEATS", font=head, fill="FFDDEBF7")
    row("AI is provisional", "m1/m2 are automated estimates and are frequently wrong, "
                            "especially on years they were trained on. The "
                            "'*_in_training_set' flag marks likely training rows — "
                            "treat those estimates as optimistic.")
    row("sex/view", "The model leaves sex as 'unknown' and view is a weak guess; "
                   "these especially need expert eyes.")
    row("2026 labels", "Photo filenames encode condo+cell+view (high confidence). "
                      "Rows sourced from the paper notebook (label_source mentions "
                      "'notebook') are a 6/5 session with NO photo — thumbnail blank "
                      "by design. A few 2026 cells are flagged in notes for review.")
    row("blank days_until_molt", "Left blank when the capture date or molt date "
                                "could not be recovered from the source — not an error.")
    row("thumbnails", "The thumbnail column (A) and the header row are frozen, so the "
                     "photo stays visible while you scroll right to the expert_ columns. "
                     "Embedded thumbnails render in Excel; if this file was imported into "
                     "Google Sheets they may not appear — use the image_relpath column / "
                     "the shared image folder to view the full photo.")
    ws.sheet_view.showGridLines = False


def write_workbook(sheet_frames: Dict[str, pd.DataFrame],
                   models: Sequence[ModelConfig],
                   out_path: Path) -> None:
    """Write the full workbook.

    Args:
        sheet_frames: ordered mapping ``sheet_name -> DataFrame`` where each
            frame has all ``DATA_KEYS`` columns plus ``abs_path``.
        models: model configs, for provenance comments.
        out_path: destination .xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)  # drop the default sheet
    _write_readme(wb, models)  # first tab: reviewer instructions
    for sheet_name, df in sheet_frames.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel tab name limit
        # Ensure every schema column exists even if the frame is empty.
        for key in DATA_KEYS:
            if key not in df.columns:
                df[key] = ""
        if "abs_path" not in df.columns:
            df["abs_path"] = ""
        _write_sheet(ws, df, models)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
