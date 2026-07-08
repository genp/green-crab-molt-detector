"""
Raw-only extractor for the folder-organized datasets (2016 / 2017 / 2018).

Every label is derived directly from raw data — folder/file names and the raw
per-year Excel workbooks. It does NOT read anything from ``data/processed``.

Per image, labels come from:
1. Ancestor folder names: crab id (``F1``/``M3``), sex (F/M or "male"/"female"),
   capture date (e.g. ``8:26``), molt date (``(molted 9:23)``), outcome
   (died/molted/thrown back).
2. For 2016 only, the authoritative raw workbook
   ``NH Green Crab Project 2016/Green Crabs September 2016.xlsx`` ("General Data"
   sheet: per-crab molt date + carapace width). When present, its molt date and
   carapace override the folder-parsed values and ``days_until_molt`` is
   recomputed from the folder capture date.

2018's workbook holds only aggregate daily molt/mortality counts (no per-crab
per-image molt date), and 2017's ``molted crabs.xlsx`` is a free-form grid, so
those two years rely on folder-name parsing. Images with no derivable label
still get a row (blank ground truth) so the sheet is a full inventory.
"""

from __future__ import annotations

from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Dict, Optional, Tuple

from ..config import RAW_DIR, YearConfig
from ..records import ImageRecord, normalize_str
from .base import (
    days_between,
    fmt_date,
    iter_images,
    make_relpath,
    parse_crab_id,
    parse_date,
    parse_outcome,
    parse_sex_from_text,
)

try:
    import openpyxl
    _HAVE_OPENPYXL = True
except Exception:  # pragma: no cover
    _HAVE_OPENPYXL = False

WORKBOOK_2016 = "NH Green Crab Project 2016/Green Crabs September 2016.xlsx"


@lru_cache(maxsize=1)
def _crab_table_2016() -> Dict[str, Dict[str, str]]:
    """Return ``{crab_id: {molt_date, carapace_width_mm}}`` from the 2016 xlsx."""
    path = RAW_DIR / WORKBOOK_2016
    if not (_HAVE_OPENPYXL and path.exists()):
        return {}
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)
    if "General Data" not in wb.sheetnames:
        return {}
    ws = wb["General Data"]
    out: Dict[str, Dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        crab = normalize_str(row[0]).upper().replace(" ", "")
        if not crab or not crab[0] in "FM":
            continue
        carapace = normalize_str(row[1]) if len(row) > 1 else ""
        molt = row[4] if len(row) > 4 else None
        molt_date = molt.strftime("%Y-%m-%d") if isinstance(molt, datetime) else ""
        out[crab] = {"molt_date": molt_date, "carapace_width_mm": carapace}
    return out


def _parse_from_folders(img_path: Path, root: Path, default_year: int) -> Dict[str, str]:
    """Derive labels by walking ancestor folder names from leaf upward."""
    rel_parts = img_path.relative_to(root).parts
    crab_id = sex = outcome = ""
    capture: Optional[datetime] = None
    molt: Optional[datetime] = None

    for part in rel_parts:
        crab_id = crab_id or parse_crab_id(part)
        sex = sex or parse_sex_from_text(part)
        outcome = outcome or parse_outcome(part)
        if "molted" in part.lower() and molt is None:
            molt = parse_date(part, default_year)

    for part in reversed(rel_parts):
        if "molted" in part.lower():
            continue
        dt = parse_date(part, default_year)
        if dt is not None:
            capture = dt
            break
    if capture is None:
        capture = parse_date(img_path.stem, default_year)

    return {
        "crab_id": crab_id,
        "sex": sex,
        "capture": capture,
        "molt": molt,
        "outcome": outcome,
    }


def extract(cfg: YearConfig) -> list[ImageRecord]:
    assert cfg.raw_subdir is not None
    root = RAW_DIR / cfg.raw_subdir
    if not root.exists():
        return []
    default_year = int(cfg.sheet_name) if cfg.sheet_name.isdigit() else 2016
    crab_table = _crab_table_2016() if cfg.sheet_name == "2016" else {}

    records = []
    for img in iter_images(root):
        parsed = _parse_from_folders(img, root, default_year)
        crab_id = parsed["crab_id"]
        sex = parsed["sex"]
        capture = parsed["capture"]
        molt = parsed["molt"]
        carapace = ""
        label_source = "raw folder names"
        confidence = "low"

        # 2016: enrich with the authoritative raw workbook.
        enrich = crab_table.get((crab_id or "").upper())
        if enrich:
            if enrich["molt_date"]:
                molt = datetime.strptime(enrich["molt_date"], "%Y-%m-%d")
            carapace = enrich["carapace_width_mm"]
            label_source = "raw folders + Green Crabs September 2016.xlsx"
            confidence = "medium"

        days = days_between(capture, molt)
        records.append(ImageRecord(
            year=cfg.sheet_name,
            dataset=cfg.raw_subdir,
            image_relpath=make_relpath(img),
            abs_path=str(img),
            crab_id=crab_id or img.parent.name,
            session_id=f"{cfg.sheet_name}::{fmt_date(capture) or 'unknown'}",
            original_filename=img.name,
            capture_date=fmt_date(capture),
            molt_date=fmt_date(molt),
            days_until_molt=days,
            sex=sex,
            carapace_width_mm=carapace,
            outcome=parsed["outcome"],
            label_source=label_source,
            label_confidence=confidence if (days or molt) else ("low" if crab_id else "unknown"),
            is_in_situ="false",
        ))
    return records
