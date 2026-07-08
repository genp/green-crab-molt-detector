"""
Lightweight smoke tests for the label_sheet package.

These avoid ML (no torch): they exercise the schema, records, extractors, and the
XLSX writer on a tiny synthetic sample. Run with:

    venv/bin/python -m pytest tests/test_label_sheet.py -q

or directly:

    venv/bin/python tests/test_label_sheet.py
"""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import pandas as pd

from label_sheet.config import MODELS
from label_sheet.records import ImageRecord, records_to_frame
from label_sheet.schema import COLUMN_KEYS, DATA_KEYS, COLUMNS
from label_sheet.xlsx_writer import write_workbook


def test_schema_unique_and_ordered():
    assert COLUMN_KEYS[0] == "thumbnail", "thumbnail must be column A"
    assert len(COLUMN_KEYS) == len(set(COLUMN_KEYS)), "duplicate column keys"
    # Every model group has the required prefixes.
    for mid in ("m1", "m2"):
        assert any(k.startswith(f"{mid}_days_to_molt") for k in COLUMN_KEYS)


def test_record_roundtrip_has_all_keys():
    rec = ImageRecord(year="2016", dataset="d", image_relpath="a/b.jpg",
                      crab_id="F2", days_until_molt="5")
    row = rec.to_row()
    assert set(row.keys()) == set(DATA_KEYS)
    assert row["crab_id"] == "F2"
    assert row["days_until_molt"] == "5"
    # Unset model/expert columns default to "".
    assert row["m1_days_to_molt"] == ""
    assert row["expert_status"] == ""


def test_records_frame_carries_abs_path():
    recs = [ImageRecord(year="2016", dataset="d", image_relpath="a.jpg", abs_path="/tmp/a.jpg")]
    frame = records_to_frame(recs)
    assert "abs_path" in frame.columns
    assert frame.iloc[0]["abs_path"] == "/tmp/a.jpg"


def test_writer_builds_workbook(tmp_path):
    import openpyxl
    frame = records_to_frame([
        ImageRecord(year="2016", dataset="d", image_relpath="a.jpg", crab_id="F1"),
        ImageRecord(year="2016", dataset="d", image_relpath="b.jpg", crab_id="F2"),
    ])
    out = tmp_path / "wb.xlsx"
    write_workbook({"2016": frame, "empty_year": frame.iloc[0:0]}, MODELS, out)
    wb = openpyxl.load_workbook(out)
    # A README instructions tab is prepended, then one tab per year.
    assert wb.sheetnames == ["README", "2016", "empty_year"]
    assert wb["README"].max_row > 5
    ws = wb["2016"]
    # Header row present with all columns; both years share identical columns.
    assert ws.max_column == len(COLUMNS)
    assert [ws.cell(1, i + 1).value for i in range(len(COLUMNS))] == [c.header for c in COLUMNS]
    # Model provenance comment exists on a model-group header.
    assert any(ws.cell(1, c).comment for c in range(1, ws.max_column + 1))
    # Empty year still has the identical header layout.
    ws2 = wb["empty_year"]
    assert ws2.max_column == len(COLUMNS)


if __name__ == "__main__":
    import tempfile
    test_schema_unique_and_ordered()
    test_record_roundtrip_has_all_keys()
    test_records_frame_carries_abs_path()
    with tempfile.TemporaryDirectory() as d:
        test_writer_builds_workbook(Path(d))
    print("all smoke tests passed")
